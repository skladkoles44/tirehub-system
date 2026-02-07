#!/usr/bin/env python3
from __future__ import annotations
import argparse, json, os, re, sys, time
from pathlib import Path
from typing import Any, Dict, List, Tuple
from openpyxl import load_workbook

def norm(x: Any) -> str:
    s = "" if x is None else str(x)
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    return s

def parse_qty(v: Any) -> Tuple[str, List[str]]:
    s0 = norm(v)
    s = s0.lower()
    if s == "":
        return "", ["missing_qty"]
    m = re.search(r"(\d+)", s)
    if not m:
        return "", ["qty_non_numeric"]
    q = m.group(1)
    flags: List[str] = []
    if not s.isdigit():
        flags.append("qty_approximated")
    return q, flags

def parse_price(v: Any) -> Tuple[str, List[str]]:
    s = norm(v)
    if s == "":
        return "", ["missing_price"]
    try:
        f = float(s.replace(",", "."))
        if f <= 0:
            return "", ["bad_price"]
        if abs(f - round(f)) < 1e-9:
            return str(int(round(f))), []
        return str(f), []
    except Exception:
        return "", ["bad_price"]

def atomic_write_text(path: Path, text: str) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text, encoding="utf-8")
    tmp.replace(path)

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--run-id", required=True)
    ap.add_argument("--sheet", default="Шины")
    ap.add_argument("--out", required=True)
    ap.add_argument("--stats-out", required=True)
    ap.add_argument("--progress-every", type=int, default=200)
    args = ap.parse_args()

    FILE = args.file
    RUN_ID = args.run_id
    SHEET = args.sheet

    # fixed mapping for Centrshin "Шины" sheet (from your inventory)
    COL_NAME = 1
    COL_SKU = 2
    COL_BRAND = 3
    COL_QTY = 12
    COL_PRICE = 13

    out_path = Path(args.out)
    st_path = Path(args.stats_out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    st_path.parent.mkdir(parents=True, exist_ok=True)

    stats: Dict[str, Any] = {
        "supplier_id": "centrshin",
        "parser_id": "centrshin_shiny_xlsx_v1",
        "parser_version": "1.0",
        "run_id": RUN_ID,
        "file": FILE,
        "sheet": SHEET,
        "lines": 0,
        "bad_json": 0,
        "bad_price": 0,
        "skipped_qty_empty": 0,
        "skipped_qty_invalid": 0,
        "missing_sku": 0,
        "flags_counts": {},
    }

    wb = load_workbook(FILE, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"sheet not found: {SHEET}. available={wb.sheetnames}")
    ws = wb[SHEET]

    # stream write (atomic rename at end)
    tmp_out = out_path.with_suffix(out_path.suffix + ".tmp")
    f = tmp_out.open("w", encoding="utf-8")

    t0 = time.time()
    try:
        max_row = ws.max_row or 0
        for r in range(2, max_row + 1):
            if args.progress_every > 0 and (r % args.progress_every == 0):
                dt = time.time() - t0
                print(f"progress: row={r}/{max_row} lines={stats['lines']} dt={dt:.1f}s", flush=True)

            sku = norm(ws.cell(row=r, column=COL_SKU).value)
            if sku == "":
                stats["missing_sku"] += 1
                continue

            price_s, pf = parse_price(ws.cell(row=r, column=COL_PRICE).value)
            if ("missing_price" in pf) or ("bad_price" in pf) or (price_s == ""):
                stats["bad_price"] += 1
                continue

            qty_cell = ws.cell(row=r, column=COL_QTY).value
            qty_s, qf = parse_qty(qty_cell)
            if qty_s == "":
                raw = norm(qty_cell)
                if raw == "":
                    stats["skipped_qty_empty"] += 1
                else:
                    stats["skipped_qty_invalid"] += 1
                continue
            try:
                if int(qty_s) <= 0:
                    stats["skipped_qty_invalid"] += 1
                    continue
            except Exception:
                stats["skipped_qty_invalid"] += 1
                continue

            flags = pf + qf
            for fl in flags:
                stats["flags_counts"][fl] = stats["flags_counts"].get(fl, 0) + 1

            rec = {
                "supplier_id": "centrshin",
                "parser_id": "centrshin_shiny_xlsx_v1",
                "parser_version": "1.0",
                "run_id": RUN_ID,
                "quality_flags": flags,
                "raw": {
                    "supplier_warehouse_name": "центршин",
                    "sku_candidate_key": sku,
                    "price": price_s,
                    "qty": qty_s,
                    "currency": "RUB",
                    "brand_raw": norm(ws.cell(row=r, column=COL_BRAND).value),
                    "name_raw": norm(ws.cell(row=r, column=COL_NAME).value),
                },
                "_meta": {"source_row_1based": r, "passthrough": {"sheet": SHEET}},
            }

            f.write(json.dumps(rec, ensure_ascii=False, separators=(",", ":")) + "\n")
            stats["lines"] += 1
    finally:
        f.close()

    tmp_out.replace(out_path)
    atomic_write_text(st_path, json.dumps(stats, ensure_ascii=False, indent=2))
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
