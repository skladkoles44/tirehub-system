#!/usr/bin/env python3
# Fast diagnostic for Brinex XLSX:
# - read_only + iter_rows(values_only=True)
# - finds header row, skips numbering row
# - samples first N data rows
# - anomaly scan with capped rows (default 3000) to avoid long runs on phone

import sys
from pathlib import Path
from openpyxl import load_workbook

NEEDED_HEADERS = {
    "Код товара (goods_id)",
    "Номенклатура",
    "Код товара (product_id)",
    "Артикул",
    "Вид товара",
    "Розница",
    "Остаток общий",
    "Склад",
    "Цена",
    "Остаток на складе",
    "Производитель",
    "Ширина",
    "Диаметр",
    "Код производителя",
    "Модель",
}

def norm(x):
    if x is None:
        return ""
    return str(x).strip()

def is_numbering_row(vals):
    xs = [norm(v) for v in vals]
    xs = [x for x in xs if x]
    if not xs:
        return False
    ok = 0
    for i, x in enumerate(xs[:50], start=1):
        if x == str(i):
            ok += 1
        else:
            break
    need = min(5, len(xs))
    return ok >= need

def parse_qty(x):
    s = norm(x)
    if s == "":
        return None, "qty_empty"
    if s.startswith(">"):
        return s, "qty_gt"
    try:
        if "," in s or "." in s:
            return float(s.replace(",", ".")), None
        return int(s), None
    except Exception:
        return s, "qty_nonnumeric"

def parse_price(x):
    s = norm(x)
    if s == "":
        return None, "price_empty"
    try:
        return float(s.replace(" ", "").replace(",", ".")), None
    except Exception:
        return s, "price_nonnumeric"

def main():
    if len(sys.argv) < 3:
        print("USAGE: diag_brinex_xlsx_v3_fast.py <xlsx_path> <sheet_name> [max_anom_rows]")
        sys.exit(2)

    xlsx = Path(sys.argv[1])
    sheet = sys.argv[2]
    max_anom_rows = int(sys.argv[3]) if len(sys.argv) >= 4 else 3000

    if not xlsx.exists():
        print("NOT FOUND:", xlsx)
        sys.exit(1)

    wb = load_workbook(xlsx, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        print("SHEETS:")
        for s in wb.sheetnames:
            print(" -", s)
        print("SHEET NOT FOUND:", sheet)
        sys.exit(1)

    ws = wb[sheet]

    # read first 120 rows to detect header
    first_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        first_rows.append((i, row))
        if i >= 120:
            break

    best_hits, hdr_row, hdr_vals = 0, None, None
    for i, row in first_rows:
        vals = [norm(v) for v in row]
        hits = sum(1 for v in vals if v in NEEDED_HEADERS)
        if hits > best_hits:
            best_hits, hdr_row, hdr_vals = hits, i, vals

    print("=== FILE ===")
    print(str(xlsx))
    print("=== SHEET ===")
    print(sheet)
    print("=== DIM ===")
    print(f"rows={ws.max_row} cols={ws.max_column}")
    print("=== HEADER DETECT (FAST) ===")
    print(f"best_hits={best_hits} header_row={hdr_row}")

    if hdr_row is None or best_hits < 6:
        print("HEADER NOT DETECTED (hits<6). STOP.")
        sys.exit(1)

    # build col index from header row
    colmap = {}
    for idx, v in enumerate(hdr_vals, start=1):
        if v:
            colmap[v] = idx

    print("=== HEADER VALUES (non-empty) ===")
    for idx, v in enumerate(hdr_vals, start=1):
        if v:
            print(f"c{idx}: {v}")

    # determine data_start
    data_start = hdr_row + 1
    # get row values for data_start from cached first_rows if possible
    row_after = None
    for i, row in first_rows:
        if i == data_start:
            row_after = row
            break
    if row_after is None:
        # read that specific row by iterating (still ok once)
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == data_start:
                row_after = row
                break

    if row_after is not None and is_numbering_row(row_after):
        data_start += 1

    print("=== DATA START ===")
    print(f"data_start_row={data_start}")

    required = ["Код товара (product_id)", "Артикул", "Склад", "Цена", "Остаток на складе"]
    miss = [r for r in required if r not in colmap]
    print("=== REQUIRED COLS CHECK ===")
    print("OK" if not miss else ("MISSING: " + ", ".join(miss)))

    def col(row, name):
        c = colmap.get(name)
        if not c:
            return None
        if c-1 < 0 or c-1 >= len(row):
            return None
        return row[c-1]

    # iterate from start; sample first 20 rows; anomaly scan capped
    print("=== SAMPLE (first 20 non-empty data rows) ===")
    sample_n = 0
    anom_hits = 0
    scanned = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < data_start:
            continue

        prod_id = col(row, "Код товара (product_id)")
        art = col(row, "Артикул")
        wh = col(row, "Склад")
        price = col(row, "Цена")
        qty = col(row, "Остаток на складе")
        name = col(row, "Номенклатура")

        if prod_id is None and art is None and wh is None and price is None and qty is None and name is None:
            continue

        qv, qflag = parse_qty(qty)
        pv, pflag = parse_price(price)
        flags = ",".join([f for f in [qflag, pflag] if f]) or "-"

        if sample_n < 20:
            print(f"r{i}: product_id={prod_id!r} article={art!r} wh={wh!r} price={price!r} qty={qty!r} flags={flags} name={norm(name)[:60]!r}")
            sample_n += 1

        # anomaly scan (cap rows)
        scanned += 1
        if scanned <= max_anom_rows and (qflag or pflag):
            reasons = ",".join([f for f in [qflag, pflag] if f])
            print(f"ANOM r{i}: reasons={reasons} product_id={prod_id!r} article={art!r} wh={wh!r} price={price!r} qty={qty!r}")
            anom_hits += 1
            if anom_hits >= 50:
                # enough
                break

        if scanned >= max_anom_rows and anom_hits >= 0 and sample_n >= 20:
            # stop early: we only needed a cap scan on phone
            break

    print("=== ANOMALY SCAN RESULT ===")
    print(f"scanned_rows={scanned} anomaly_hits={anom_hits} cap={max_anom_rows}")

if __name__ == "__main__":
    main()
