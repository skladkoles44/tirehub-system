#!/usr/bin/env python3
# Brinex XLSX emitter for ANY sheet with Brinex headers.
# Produces canonical good.ndjson + stats.json compatible with tirehub_ingest_v1.
import json, sys, hashlib, datetime, secrets, re
from pathlib import Path
from openpyxl import load_workbook

SUPPLIER_ID = "brinex"
EMITTER_VERSION = "brinex_xlsx_sheet_v1"
NDJSON_CONTRACT_VERSION = "good_ndjson_v1"
MAPPING_VERSION = "1.0"

NEEDED_HEADERS = {
    "Код товара (goods_id)","Номенклатура","Код товара (product_id)","Артикул","Вид товара","Розница",
    "Остаток общий","Склад","Цена","Остаток на складе","Производитель","Ширина","Высота","Диаметр",
    "Сезонность","Индекс нагрузки/Индекс нагрузки сдвоенной шины","Индекс скорости","Run Flat",
    "Код производителя","Модель","Шипы",
}

def utc_now_rfc3339z():
    return datetime.datetime.now(datetime.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

def norm(x):
    if x is None: return ""
    return str(x).strip()

def slug_sheet(s: str) -> str:
    import hashlib
    s = (s or "").strip()
    h = hashlib.sha1(s.encode("utf-8")).hexdigest()[:10]
    return f"sheet_{h}"

def is_numbering_row(vals):
    xs = [norm(v) for v in vals]
    xs = [x for x in xs if x]
    if not xs: return False
    ok = 0
    for i, x in enumerate(xs[:50], start=1):
        if x == str(i): ok += 1
        else: break
    need = min(5, len(xs))
    return ok >= need

def parse_qty(x):
    s = norm(x)
    if s == "": return None, "qty_empty"
    if s.startswith(">"): return s, "qty_gt"
    try:
        if "," in s or "." in s: return float(s.replace(",", ".")), None
        return int(s), None
    except Exception:
        return s, "qty_nonnumeric"

def parse_price(x):
    s = norm(x)
    if s == "":
        return None, "price_empty"
    try:
        v = float(s.replace(" ", "").replace(",", "."))
        if v.is_integer():
            return int(v), None
        return int(round(v)), "price_coerced"
    except Exception:
        return s, "price_nonnumeric"

def sha256_hex(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()

def main():
    if len(sys.argv) < 4:
        print("USAGE: emit_brinex_xlsx_sheet_v1.py <xlsx_path> <sheet_name> <out_dir>")
        sys.exit(2)

    xlsx_path = Path(sys.argv[1])
    sheet_name = sys.argv[2]
    out_dir = Path(sys.argv[3])

    if not xlsx_path.exists():
        print(f"NOT FOUND: {xlsx_path}")
        sys.exit(1)

    out_dir.mkdir(parents=True, exist_ok=True)

    sheet_slug = slug_sheet(sheet_name)
    parser_id = f"brinex__{sheet_slug}__xlsx_v1"

    run_id = f"{SUPPLIER_ID}_{datetime.datetime.now(datetime.timezone.utc).strftime('%Y%m%dT%H%M%SZ')}_{secrets.token_hex(3)}"
    effective_at = utc_now_rfc3339z()

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        print("SHEETS:")
        for s in wb.sheetnames: print(" -", s)
        print(f"SHEET NOT FOUND: {sheet_name}")
        sys.exit(1)

    ws = wb[sheet_name]

    # detect header row within first 120 rows
    first_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        first_rows.append((i, row))
        if i >= 120: break

    best_hits, hdr_row, hdr_vals = 0, None, None
    for i, row in first_rows:
        vals = [norm(v) for v in row]
        hits = sum(1 for v in vals if v in NEEDED_HEADERS)
        if hits > best_hits:
            best_hits, hdr_row, hdr_vals = hits, i, vals

    if hdr_row is None or best_hits < 6:
        print(f"SKIP: header not detected (best_hits={best_hits}) sheet={sheet_name!r}")
        sys.exit(3)

    colmap = {}
    for idx, v in enumerate(hdr_vals, start=1):
        if v: colmap[v] = idx

    data_start = hdr_row + 1
    row_after = None
    for i, row in first_rows:
        if i == data_start:
            row_after = row
            break
    if row_after is not None and is_numbering_row(row_after):
        data_start += 1

    mapping = {
        "product_id": "Код товара (product_id)",
        "article": "Артикул",
        "warehouse": "Склад",
        "price": "Цена",
        "qty": "Остаток на складе",
        "name": "Номенклатура",
        "brand": "Производитель",
        "width": "Ширина",
        "height": "Высота",
        "diameter": "Диаметр",
        "season": "Сезонность",
        "load_index": "Индекс нагрузки/Индекс нагрузки сдвоенной шины",
        "speed_index": "Индекс скорости",
        "runflat": "Run Flat",
        "mfg_code": "Код производителя",
        "model": "Модель",
        "studded": "Шипы",
        "goods_id": "Код товара (goods_id)",
        "retail": "Розница",
        "stock_total": "Остаток общий",
        "item_type": "Вид товара",
    }
    mapping_hash = sha256_hex(json.dumps(mapping, ensure_ascii=False, sort_keys=True))

    good_path = out_dir / "good.ndjson"
    stats_path = out_dir / "stats.json"

    def col(row, header):
        c = colmap.get(header)
        if not c: return None
        j = c - 1
        if j < 0 or j >= len(row): return None
        return row[j]

    good_rows = 0
    bad_rows = 0
    flags_counts = {}
    source_rows_read = 0

    with good_path.open("w", encoding="utf-8") as fout:
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i < data_start: continue
            if not any(v is not None and norm(v) != "" for v in row): continue

            source_rows_read += 1

            product_id = col(row, "Код товара (product_id)")
            article = col(row, "Артикул")
            warehouse = col(row, "Склад")
            price_raw = col(row, "Цена")
            qty_raw = col(row, "Остаток на складе")

            name = col(row, "Номенклатура")
            brand = col(row, "Производитель")
            width = col(row, "Ширина")
            height = col(row, "Высота")
            diameter = col(row, "Диаметр")
            season = col(row, "Сезонность")
            load_index = col(row, "Индекс нагрузки/Индекс нагрузки сдвоенной шины")
            speed_index = col(row, "Индекс скорости")
            runflat = col(row, "Run Flat")
            mfg_code = col(row, "Код производителя")
            model = col(row, "Модель")
            studded = col(row, "Шипы")

            qty, qflag = parse_qty(qty_raw)
            price, pflag = parse_price(price_raw)

            row_flags = []
            for fl in [qflag, pflag]:
                if fl:
                    row_flags.append(fl)
                    flags_counts[fl] = flags_counts.get(fl, 0) + 1

            sku_candidate = product_id if product_id not in (None, "") else article
            if sku_candidate in (None, ""): row_flags.append("sku_missing")
            if warehouse in (None, ""): row_flags.append("warehouse_missing")
            if qty is None: row_flags.append("qty_missing")
            if price is None: row_flags.append("price_missing")

            if row_flags:
                bad_rows += 1
                for fl in row_flags:
                    flags_counts[fl] = flags_counts.get(fl, 0) + 1
                continue

            parsed = {
                "product_id": product_id,
                "article": article,
                "warehouse": warehouse,
                "qty": qty,
                "price": price,
                "name": norm(name) or None,
                "brand": norm(brand) or None,
                "width": norm(width) or None,
                "height": norm(height) or None,
                "diameter": norm(diameter) or None,
                "season": norm(season) or None,
                "load_index": norm(load_index) or None,
                "speed_index": norm(speed_index) or None,
                "runflat": norm(runflat) or None,
                "mfg_code": norm(mfg_code) or None,
                "model": norm(model) or None,
                "studded": norm(studded) or None,
            }

            raw = {
                "source": {
                    "source_path": str(xlsx_path) + "::" + sheet_name,  # invalid path -> ingest will fallback to good_path
                    "sheet": sheet_name,
                    "source_row_number": i,
                    "header_row": hdr_row,
                    "data_start_row": data_start,
                },
                "cells": {
                    "product_id": product_id,
                    "article": article,
                    "warehouse": warehouse,
                    "qty_raw": qty_raw,
                    "price_raw": price_raw,
                }
            }

            good_obj = {
                "supplier_id": SUPPLIER_ID,
                "sku_candidate_key": "product_id",
                "sku_candidate": str(product_id) if product_id not in (None, "") else str(article),
                "parser_id": parser_id,
                "run_id": run_id,
                "effective_at": effective_at,
                "mapping_version": MAPPING_VERSION,
                "mapping_hash": mapping_hash,
                "ndjson_contract_version": NDJSON_CONTRACT_VERSION,
                "emitter_version": EMITTER_VERSION,
                "raw": raw,
                "parsed": parsed,
                "quality_flags": [],
                "_meta": {
                    "sheet": sheet_name,
                    "sheet_slug": sheet_slug,
                    "header_row": hdr_row,
                    "data_start_row": data_start,
                    "source_row_number": i,
                }
            }

            fout.write(json.dumps(good_obj, ensure_ascii=False) + "\n")
            good_rows += 1

    stats = {
        "supplier_id": SUPPLIER_ID,
        "parser_id": parser_id,
        "sheet": sheet_name,
        "run_id": run_id,
        "effective_at": effective_at,
        "mapping_version": MAPPING_VERSION,
        "mapping_hash": mapping_hash,
        "ndjson_contract_version": NDJSON_CONTRACT_VERSION,
        "emitter_version": EMITTER_VERSION,
        "good_rows": good_rows,
        "bad_rows": bad_rows,
        "exploded_lines": good_rows,
        "source_rows_read": source_rows_read,
        "flags_counts": flags_counts,
        "source_path": str(xlsx_path) + "::" + sheet_name,  # invalid path -> ingest will fallback to good_path
        "good_path": str(good_path),
    }
    stats_path.write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print("OK")
    print("sheet=", sheet_name)
    print("parser_id=", parser_id)
    print("run_id=", run_id)
    print("good_rows=", good_rows)
    print("bad_rows=", bad_rows)
    print("source_rows_read=", source_rows_read)

if __name__ == "__main__":
    main()
