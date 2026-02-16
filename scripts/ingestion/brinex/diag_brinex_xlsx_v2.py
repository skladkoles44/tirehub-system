#!/usr/bin/env python3
# Brinex XLSX diagnostic v2:
# - exact header match (no substring false positives)
# - numbering row detection works even with fewer cols
# - prints sample + scans for anomalies in qty/price
import sys
from pathlib import Path
from openpyxl import load_workbook

NEEDED_HEADERS = [
    "Код товара (goods_id)",
    "Номенклатура",
    "Код товара (product_id)",
    "Артикул",
    "Вид товара",
    "Розница",
    "Остаток общий",
    "Склад",
    "Цена",
    "Остаток на складе",
    "Производитель",
    "Ширина",
    "Диаметр",
    "Код производителя",
    "Модель",
]

def norm(s):
    if s is None:
        return ""
    return str(s).strip()

def is_intlike(x: str) -> bool:
    if x == "":
        return False
    if x.startswith("+"):
        x = x[1:]
    return x.isdigit()

def is_numbering_row(row_vals):
    # Detect 1..N sequence; accept >= min(5, nonempty_count) consecutive
    xs = [norm(x) for x in row_vals]
    xs = [x for x in xs if x != ""]
    if not xs:
        return False
    # count consecutive "1","2","3"... from start
    ok = 0
    for i, x in enumerate(xs[:50], start=1):
        if x == str(i):
            ok += 1
        else:
            break
    need = min(5, len(xs))
    return ok >= need

def find_header_row(ws, max_scan=120):
    # Exact match hits: header cell value must equal one of needed headers
    best = (0, None, None)  # (hits, row_idx, row_vals_norm)
    for r in range(1, min(ws.max_row, max_scan) + 1):
        vals = [norm(c.value) for c in ws[r]]
        hits = sum(1 for v in vals if v in NEEDED_HEADERS)
        if hits > best[0]:
            best = (hits, r, vals)
    return best

def build_colmap(header_vals):
    colmap = {}
    for idx, v in enumerate(header_vals, start=1):
        if v:
            colmap[v] = idx
    return colmap

def parse_qty(x):
    s = norm(x)
    if s == "":
        return None, "qty_empty"
    if s.startswith(">"):
        # keep as special marker; caller can convert to numeric policy later
        return s, "qty_gt"
    # allow ints or floats
    try:
        if "." in s or "," in s:
            s2 = s.replace(",", ".")
            return float(s2), None
        return int(s), None
    except Exception:
        return s, "qty_nonnumeric"

def parse_price(x):
    s = norm(x)
    if s == "":
        return None, "price_empty"
    try:
        s2 = s.replace(" ", "").replace(",", ".")
        return float(s2), None
    except Exception:
        return s, "price_nonnumeric"

def main():
    if len(sys.argv) < 3:
        print("USAGE: diag_brinex_xlsx_v2.py <xlsx_path> <sheet_name>")
        sys.exit(2)

    xlsx_path = Path(sys.argv[1])
    sheet = sys.argv[2]

    if not xlsx_path.exists():
        print(f"NOT FOUND: {xlsx_path}")
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        print("SHEETS:")
        for s in wb.sheetnames:
            print(" -", s)
        print(f"SHEET NOT FOUND: {sheet}")
        sys.exit(1)

    ws = wb[sheet]
    hits, hdr_row, hdr_vals = find_header_row(ws)

    print("=== FILE ===")
    print(str(xlsx_path))
    print("=== SHEET ===")
    print(sheet)
    print("=== DIM ===")
    print(f"rows={ws.max_row} cols={ws.max_column}")
    print("=== HEADER DETECT (EXACT) ===")
    print(f"best_hits={hits} header_row={hdr_row}")

    if hdr_row is None or hits < 6:
        print("HEADER NOT DETECTED (hits<6). STOP.")
        sys.exit(1)

    colmap = build_colmap(hdr_vals)
    print("=== HEADER VALUES (non-empty) ===")
    for i, v in enumerate(hdr_vals, start=1):
        if v:
            print(f"c{i}: {v}")

    data_start = hdr_row + 1
    next_vals = [c.value for c in ws[data_start]]
    if is_numbering_row(next_vals):
        data_start += 1

    print("=== DATA START ===")
    print(f"data_start_row={data_start}")

    required = ["Код товара (product_id)", "Артикул", "Склад", "Цена", "Остаток на складе"]
    miss = [r for r in required if r not in colmap]
    print("=== REQUIRED COLS CHECK ===")
    print("OK" if not miss else ("MISSING: " + ", ".join(miss)))

    def get(row, name):
        c = colmap.get(name)
        if not c:
            return None
        return ws.cell(row=row, column=c).value

    # sample
    print("=== SAMPLE (first 20 non-empty data rows) ===")
    printed = 0
    for r in range(data_start, min(ws.max_row, data_start + 500) + 1):
        prod_id = get(r, "Код товара (product_id)")
        art = get(r, "Артикул")
        wh = get(r, "Склад")
        price = get(r, "Цена")
        qty = get(r, "Остаток на складе")
        name = get(r, "Номенклатура")
        if prod_id is None and art is None and wh is None and price is None and qty is None and name is None:
            continue
        qv, qflag = parse_qty(qty)
        pv, pflag = parse_price(price)
        flags = ",".join([f for f in [qflag, pflag] if f]) or "-"
        print(f"r{r}: product_id={prod_id!r} article={art!r} wh={wh!r} price={price!r} qty={qty!r} flags={flags} name={norm(name)[:60]!r}")
        printed += 1
        if printed >= 20:
            break

    # anomaly scan
    print("=== ANOMALY SCAN (first 50 hits) ===")
    hits2 = 0
    for r in range(data_start, ws.max_row + 1):
        price = get(r, "Цена")
        qty = get(r, "Остаток на складе")
        qv, qflag = parse_qty(qty)
        pv, pflag = parse_price(price)
        if qflag or pflag:
            prod_id = get(r, "Код товара (product_id)")
            art = get(r, "Артикул")
            wh = get(r, "Склад")
            reasons = ",".join([f for f in [qflag, pflag] if f])
            print(f"r{r}: reasons={reasons} product_id={prod_id!r} article={art!r} wh={wh!r} price={price!r} qty={qty!r}")
            hits2 += 1
            if hits2 >= 50:
                break
    print(f"anomaly_hits={hits2}")

if __name__ == "__main__":
    main()
