from __future__ import annotations

import argparse
import json
import re
import sys
import time
from pathlib import Path
from typing import Any, Dict, Optional

import yaml
from openpyxl import load_workbook


def _atomic_write_text(path: Path, text: str) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text, encoding="utf-8")
    tmp.replace(path)


def _as_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(" ", "").replace(",", ".")
    if s == "" or s.lower() in ("none", "null", "nan"):
        return None
    try:
        return float(s)
    except Exception:
        return None


def _as_int_qty(x: Any) -> tuple[Optional[int], Optional[int], Optional[str]]:
    if x is None:
        return (None, None, None)
    if isinstance(x, (int, float)):
        try:
            return (int(float(x)), None, None)
        except Exception:
            return (None, None, str(x))
    s = str(x).strip()
    if s == "":
        return (None, None, "")
    m = re.match(r"^>\s*(\d+)\s*$", s)
    if m:
        lb = int(m.group(1))
        return (lb, lb, s)
    try:
        return (int(float(s.replace(",", ".").replace(" ", ""))), None, s)
    except Exception:
        return (None, None, s)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--layout", required=True)      # category:<key>
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--run-id", required=True)
    ap.add_argument("--mapping", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--stats-out", required=True)
    ap.add_argument("--max-rows", type=int, default=None)
    ap.add_argument("--heartbeat", type=int, default=1000)
    args = ap.parse_args()

    if not args.layout.startswith("category:"):
        print("ERROR: layout must be category:<key>", file=sys.stderr)
        return 2

    category_key = args.layout.split(":", 1)[1]
    sheet_name = args.sheet
    t0 = time.time()

    # mapping defaults
    with open(args.mapping, "r", encoding="utf-8") as f:
        mapping = yaml.safe_load(f) or {}
    defaults = mapping.get("defaults", {})
    header_row = int(defaults.get("header_row", 6))
    data_row = int(defaults.get("data_row", header_row + 1))

    wb = load_workbook(args.file, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: sheet not found: {sheet_name}", file=sys.stderr)
        return 3
    ws = wb[sheet_name]

    # header map
    header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    col_map: Dict[str, int] = {}
    for idx, v in enumerate(header_cells):
        if isinstance(v, str) and v.strip():
            col_map[v.strip()] = idx

    def col(name: str) -> Optional[int]:
        return col_map.get(name)

    required = [
        "Код товара (goods_id)",
        "Номенклатура",
        "Артикул",
        "Цена",
        "Склад",
        "Остаток на складе",
    ]
    missing = [x for x in required if col(x) is None]
    if missing:
        print(f"ERROR: missing columns: {missing}", file=sys.stderr)
        return 5

    c_goods = col("Код товара (goods_id)")
    c_name = col("Номенклатура")
    c_article = col("Артикул")
    c_price = col("Цена")
    c_wh = col("Склад")
    c_qty = col("Остаток на складе")

    out_nd = Path(args.out)
    out_nd.parent.mkdir(parents=True, exist_ok=True)
    tmp_nd = out_nd.with_suffix(out_nd.suffix + ".tmp")

    emitted = seen = bad_price = bad_qty = 0
    max_rows = args.max_rows

    with tmp_nd.open("w", encoding="utf-8") as w:
        for idx, row in enumerate(ws.iter_rows(min_row=data_row, values_only=True), start=1):
            if max_rows and idx > max_rows:
                break

            goods_id = row[c_goods]
            if goods_id is None:
                continue
            seen += 1

            qty, _, _ = _as_int_qty(row[c_qty])
            if qty is None or qty <= 0:
                bad_qty += 1
                continue

            price = _as_float(row[c_price])
            if price is None or price <= 0:
                bad_price += 1
                continue

            rec = {
                "supplier_id": "brinex",
                "parser_id": f"brinex__{category_key}__xlsx_v1",
                "layout": f"category:{category_key}",
                "ts_ingested": time.strftime("%Y-%m-%dT%H:%M:%S"),
                "item": {
                    "supplier_item_id": str(goods_id).strip(),
                    "name": str(row[c_name]).strip() if row[c_name] else None,
                    "article": str(row[c_article]).strip() if row[c_article] else None,
                    "qty": qty,
                    "price_opt": price,
                    "currency": "RUB",
                },
                "raw": {
                    "sheet": sheet_name,
                    "warehouse": row[c_wh],
                },
            }
            w.write(json.dumps(rec, ensure_ascii=False) + "\n")
            emitted += 1

            if args.heartbeat and idx % args.heartbeat == 0:
                print(f"[{category_key}] rows={idx} seen={seen} emitted={emitted}", file=sys.stderr)

    tmp_nd.replace(out_nd)

    stats = {
        "category_key": category_key,
        "sheet_name": sheet_name,
        "rows_seen": seen,
        "rows_emitted": emitted,
        "bad_price": bad_price,
        "bad_qty": bad_qty,
        "elapsed_sec": round(time.time() - t0, 1),
    }
    _atomic_write_text(Path(args.stats_out), json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
