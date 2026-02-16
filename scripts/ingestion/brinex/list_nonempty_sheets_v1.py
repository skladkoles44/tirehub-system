#!/usr/bin/env python3
import sys
from pathlib import Path
from openpyxl import load_workbook

def has_any_value(ws, scan_rows=80):
    max_r = min(ws.max_row or 0, scan_rows)
    if max_r <= 0:
        return False
    for r in range(1, max_r + 1):
        row = ws[r]
        for c in row:
            if c.value not in (None, ""):
                return True
    return False

def main():
    if len(sys.argv) != 2:
        print("USAGE: list_nonempty_sheets_v1.py <xlsx_path>", file=sys.stderr)
        return 2
    xlsx = Path(sys.argv[1])
    if not xlsx.exists():
        print(f"XLSX NOT FOUND: {xlsx}", file=sys.stderr)
        return 1

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        # max_row иногда завышен, поэтому проверяем наличие реальных значений
        if (ws.max_row or 0) > 1 and has_any_value(ws):
            print(name)
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
