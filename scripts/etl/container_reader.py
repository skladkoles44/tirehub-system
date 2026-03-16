from pathlib import Path

class XlsSheetAdapter:
    def __init__(self, name, rows):
        self.title = name
        self._rows = rows

    @property
    def max_row(self):
        return len(self._rows)

    @property
    def max_column(self):
        return max((len(r) for r in self._rows), default=0)

    def iter_rows(self, values_only=False):
        for row in self._rows:
            tup = tuple(row)
            yield tup

    def __iter__(self):
        return self.iter_rows(values_only=True)


def container_reader(path: Path):
    ext = path.suffix.lower()

    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            yield sheet_name, wb[sheet_name]

    elif ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        for sheet in wb.sheets():
            rows = [sheet.row_values(r) for r in range(sheet.nrows)]
            yield sheet.name, XlsSheetAdapter(sheet.name, rows)

    else:
        raise RuntimeError(f"UNSUPPORTED_FILE_TYPE={ext}")
