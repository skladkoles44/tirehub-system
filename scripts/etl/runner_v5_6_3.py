#!/usr/bin/env python3
"""
runner_v5_6_3.py — с fallback для "плохих" таблиц
"""

import sys
import os
import json
import hashlib
import csv
from pathlib import Path
from datetime import datetime, timezone

import openpyxl
import xlrd
import odf.opendocument
from odf.table import Table, TableRow, TableCell
from odf.text import P

MAX_CELL = 10000
MAX_COLUMNS = 10000
MAX_JSON_BYTES = 50 * 1024 * 1024
FSYNC_EVERY = 1000


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return f"sha256:{h.hexdigest()}"


def fsync_directory(path: Path):
    try:
        fd = os.open(str(path.parent), os.O_RDONLY | os.O_DIRECTORY)
        try:
            os.fsync(fd)
        finally:
            os.close(fd)
    except Exception:
        pass


def trim_value(v):
    if isinstance(v, str) and len(v) > MAX_CELL:
        return v[:MAX_CELL]
    return v


def normalize_value(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def is_empty(row):
    for v in row:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def append_line(f, obj, i):
    f.write(json.dumps(obj, ensure_ascii=False) + "\n")
    if i % FSYNC_EVERY == 0:
        f.flush()
        os.fsync(f.fileno())


# ==================== HEADER DETECTION WITH FALLBACK ====================

def detect_header_row(ws, max_scan=30):
    """Находит строку с заголовками, fallback если не нашли"""
    best_row = None
    best_score = -1
    best_idx = 0
    
    keywords = ["артикул", "наименование", "цена", "склад", "остаток", 
                "количество", "номер", "модель", "бренд", "размер", "сезон"]
    
    for idx in range(1, min(max_scan, ws.max_row + 1)):
        row = [cell.value for cell in ws[idx]]
        text_cells = [str(x).strip() for x in row if x and str(x).strip()]
        if not text_cells:
            continue
        
        text = " ".join(text_cells).lower()
        kw_score = sum(1 for kw in keywords if kw in text)
        text_ratio = len(text_cells) / len(row)
        
        # Бонус: если следующая строка содержит числа
        has_numbers = False
        if idx + 1 <= ws.max_row:
            next_row = [cell.value for cell in ws[idx + 1]]
            num_count = sum(1 for x in next_row if isinstance(x, (int, float)) and x not in (None, ""))
            if num_count >= 3:
                has_numbers = True
        
        score = kw_score * 5 + text_ratio * 10 + (10 if has_numbers else 0)
        
        if score > best_score:
            best_score = score
            best_row = [str(x) if x else "" for x in row]
            best_idx = idx
    
    # FALLBACK: если не нашли, берём первую строку
    if best_idx == 0 and ws.max_row > 0:
        first_row = [cell.value for cell in ws[1]]
        best_row = [str(x) if x else "" for x in first_row]
        best_idx = 1
    
    return best_row, best_idx


# ==================== TABLE DETECTION WITH FALLBACK ====================

def detect_tables_with_fallback(ws):
    """Определяет таблицы, если не находит — берёт весь лист"""
    tables = []
    
    # Пытаемся найти таблицы стандартным методом
    # (здесь может быть твоя логика detect_tables)
    
    # FALLBACK: если таблиц нет, берём весь лист
    if not tables:
        tables = [{
            "start_row": 1,
            "end_row": ws.max_row or 1000
        }]
    
    return tables


# ==================== READERS ====================


def read_xlsx(p):
    """Чтение Excel: fast-path (pandas) для маленьких файлов, fallback на openpyxl"""
    
    # Fast-path для файлов < 5MB
    if p.stat().st_size < 5 * 1024 * 1024:
        try:
            import pandas as pd
            import sys
            
            # Читаем файл один раз
            xls = pd.ExcelFile(p)
            
            for sheet_name in xls.sheet_names:
                # Читаем лист с ограничением по колонкам (первые 200)
                df = xls.parse(sheet_name=sheet_name, header=None)
                if df.empty:
                    continue
                
                # Ограничиваем колонки для производительности
                if df.shape[1] > 200:
                    df = df.iloc[:, :200]
                
                # Определяем заголовки
                headers = None
                header_row_idx = 0
                keywords = ["артикул", "наименование", "цена", "склад", "остаток", 
                            "количество", "номер", "модель", "бренд", "размер"]
                
                for i in range(min(15, len(df))):
                    row = df.iloc[i].fillna("").astype(str).tolist()
                    text_cells = [x for x in row if x.strip()]
                    if not text_cells:
                        continue
                    text = " ".join(text_cells).lower()
                    kw_score = sum(1 for kw in keywords if kw in text)
                    if kw_score >= 2:
                        headers = row
                        header_row_idx = i
                        break
                
                if headers is None and len(df) > 0:
                    headers = df.iloc[0].fillna("").astype(str).tolist()
                    header_row_idx = 0
                
                for i in range(header_row_idx + 1, len(df)):
                    row = df.iloc[i].fillna("").tolist()
                    if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                        continue
                    yield sheet_name, i + 1, headers, row
            return
        except Exception as e:
            # Пишем ошибку в stderr, не ломая stdout
            import sys
            print(f"⚠️ pandas fast-path failed: {e}, fallback to openpyxl", file=sys.stderr)
    
    # Fallback: старый проверенный код
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    try:
        for s in wb.sheetnames:
            ws = wb[s]
            if ws.max_row <= 1:
                continue
            headers, header_row_idx = detect_header_row(ws)
            if not headers:
                continue
            for i in range(header_row_idx + 1, ws.max_row + 1):
                row = [cell.value for cell in ws[i]]
                if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                    continue
                yield s, i, headers, row
    finally:
        wb.close()



def read_xls(p):
    try:
        wb = xlrd.open_workbook(p, on_demand=True)
    except Exception:
        return
    
    try:
        for si, name in enumerate(wb.sheet_names()):
            sh = wb.sheet_by_index(si)
            
            # Определяем заголовки с fallback
            headers = None
            header_row_idx = 0
            
            for i in range(min(30, sh.nrows)):
                row = sh.row_values(i)
                text_cells = [str(x).strip() for x in row if x and str(x).strip()]
                if not text_cells:
                    continue
                
                text = " ".join(text_cells).lower()
                keywords = ["артикул", "наименование", "цена", "склад", "остаток"]
                if any(kw in text for kw in keywords):
                    headers = [str(x) if x else "" for x in row]
                    header_row_idx = i
                    break
            
            if not headers and sh.nrows > 0:
                headers = [str(x) if x else "" for x in sh.row_values(0)]
                header_row_idx = 0
            
            for i in range(header_row_idx + 1, sh.nrows):
                r = sh.row_values(i)
                if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
                    continue
                yield name, i + 1, headers, r
    finally:
        try:
            wb.release_resources()
        except Exception:
            pass


def read_csv(p):
    with p.open("r", encoding="utf-8-sig", errors="replace") as f:
        first = None
        ln = 0
        for ln, line in enumerate(f, 1):
            if line.strip():
                first = line.strip()
                break
        if not first:
            return
        
        d = "\t" if first.count("\t") > first.count(",") else ","
        headers = list(csv.reader([first], delimiter=d))[0]
        
        for i, r in enumerate(csv.reader(f, delimiter=d), ln + 1):
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
                continue
            yield "csv", i, headers, r


def read_json(p):
    if p.stat().st_size > MAX_JSON_BYTES:
        return
    
    with p.open("r", encoding="utf-8-sig", errors="replace") as f:
        try:
            data = json.load(f)
        except Exception:
            return
    
    if isinstance(data, dict):
        headers = list(data.keys())
        yield "json", 1, headers, list(data.values())
    
    elif isinstance(data, list) and data and isinstance(data[0], dict):
        headers = list(data[0].keys())
        for i, row in enumerate(data, 1):
            yield "json", i, headers, [row.get(h) for h in headers]


def read_ods(p):
    try:
        doc = odf.opendocument.load(str(p))
    except Exception:
        return
    
    for t in doc.getElementsByType(Table):
        name = t.getAttribute("name") or "sheet"
        headers = None
        idx = 0
        header_rows = []
        
        for r in t.getElementsByType(TableRow):
            idx += 1
            cells = []
            for c in r.getElementsByType(TableCell):
                rep = int(c.getAttribute("numbercolumnsrepeated") or 1)
                txt = ""
                ps = c.getElementsByType(P)
                if ps:
                    txt = "".join(x.firstChild.data for x in ps if x.firstChild)
                cells.extend([txt] * rep)
            
            if all(not v.strip() for v in cells if v):
                continue
            
            if headers is None:
                if len(header_rows) < 3:
                    header_rows.append(cells)
                    continue
                else:
                    # Склеиваем заголовки из накопленных строк
                    max_cols = max(len(r) for r in header_rows)
                    merged = []
                    for ci in range(max_cols):
                        parts = []
                        for row in header_rows:
                            if ci < len(row) and row[ci]:
                                parts.append(str(row[ci]).strip())
                        merged.append(" | ".join(parts) if parts else f"col_{ci}")
                    headers = merged
                    yield name, idx, headers, cells
                    continue
            
            yield name, idx, headers, cells


# ==================== CORE ====================

def run(inp: Path, out: Path, file_hash: str = None):
    # Преобразуем строки в Path
    inp = Path(inp) if not isinstance(inp, Path) else inp
    out = Path(out) if not isinstance(out, Path) else out
    h = file_hash if file_hash else sha256_file(inp)
    
    out.mkdir(parents=True, exist_ok=True)
    fpath = out / "atomic_rows.ndjson"
    
    rows = 0
    
    if inp.suffix.lower() in {".xlsx", ".xlsm"}:
        reader = read_xlsx(inp)
    elif inp.suffix.lower() in {".xls", ".xlt"}:
        reader = read_xls(inp)
    elif inp.suffix.lower() in {".csv", ".tsv"}:
        reader = read_csv(inp)
    elif inp.suffix.lower() == ".json":
        reader = read_json(inp)
    elif inp.suffix.lower() == ".ods":
        reader = read_ods(inp)
    else:
        return
    
    with fpath.open("w", encoding="utf-8") as f:
        for sheet, idx, headers, row in reader:
            cols = []
            L = min(max(len(headers), len(row)), MAX_COLUMNS)
            
            for i in range(L):
                hname = headers[i] if i < len(headers) else ""
                val = row[i] if i < len(row) else None
                
                cols.append({
                    "index": i,
                    "header": hname,
                    "value": (
                        normalize_value(trim_value(val))
                        if (
                            any(k in str(hname).lower() for k in ("остат", "qty", "stock", "цена", "price"))
                            or (
                                isinstance(val, str)
                                and val.strip().replace(",", ".").replace(".", "", 1).isdigit()
                            )
                        )
                        else trim_value(val)
                    )
                })
            
            rec = {
                "row_id": f"{h}:{sheet}:{idx}",
                "source_file": str(inp),
                "file_hash": h,
                "ingestion_id": h,
                "sheet": sheet,
                "row_index": idx,
                "columns": cols
            }
            
            rows += 1
            append_line(f, rec, rows)
        
        f.flush()
        os.fsync(f.fileno())
    
    fsync_directory(fpath)
    print(f"  📊 {rows} rows extracted from {inp.name}")


def main():
    if len(sys.argv) < 3:
        print("usage: runner_v5_6_3.py <input> <out_dir>")
        sys.exit(1)
    
    run(Path(sys.argv[1]), Path(sys.argv[2]))


if __name__ == "__main__":
    main()
